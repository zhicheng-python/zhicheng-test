#encoding:utf-8
#author:wangzhicheng
#time:2019/11/28 11:14
#file:handle_excel.py


from openpyxl import load_workbook
from Test_scripts.handle_yaml import do_yaml
from Test_scripts.handle_path import do_path
from openpyxl.styles import PatternFill,colors


class DataObj:
    pass

class ReadExcel:

    def __init__(self,sheet_name=do_yaml.read_yaml("excel","sheet_name"),path=do_path.data_path):
        self.path=path
        self.sheet_name=sheet_name

    #打开文件
    def open_excel(self):
        #创建文件对象
        self.wb = load_workbook(self.path)
        #创建表单对象，指定表单
        self.sheet=self.wb[self.sheet_name]

    #获取表头
    def get_title(self):
        #调用实例方法，打开文件
        self.open_excel()

        #sheet.rows 按行获取所有数据，取第一行数据索引为0
        title=[i.value for i in list(self.sheet.rows)[0]]#列表推导式

        #关闭文件
        self.wb.close()
        return title

    #获取数据（字典数据格式）
    def get_data(self):
        title=self.get_title()
        self.open_excel()

        #sheet.rows 按行获取所有数据，取第一行数据以外的所有行索引为[1:]
        #zip聚合打包 title 和 每行每单元格的值，转换成字典
        datas=[dict(zip(title,[j.value for j in i]))for i in list(self.sheet.rows)[1:]] #列表推导式

        self.wb.close()
        return datas


    #获取数据 （[obj,obj,obj....]格式）
    def get_data_obj(self):

        title=self.get_title()
        self.open_excel()

        datas=[]

        for i in list(self.sheet.rows)[1:]:
            data=[j.value for j in i]
            data_obj=DataObj()

            for k in zip(title,data):
                setattr(data_obj,k[0],k[1])
            datas.append(data_obj)
        return datas

    #写入数据
    def write_data(self,row,column,value):

        self.open_excel()

        # 将value 值写入到指定单元格中
        self.sheet.cell(row,column,value)

        #写入之后保存文件
        self.wb.save(self.path)
        self.wb.close()

    #设置背景颜色
    def write_color(self):

        title=self.get_title()
        self.open_excel()

        #设置背景颜色
        fill_1=PatternFill(patternType="solid",fgColor=colors.RED)
        fill_2=PatternFill(patternType="solid",fgColor=colors.GREEN)

        #遍历 result 这一列，单元格的值==pass则用fill_1，==fail用fill_2
        for i in list(self.sheet.columns)[title.index("result")]:
            if i.value =="PASS":
                i.fill=fill_2
            if i.value=="FAIL":
                i.fill=fill_1

        self.wb.save(self.path)

        self.wb.close()




do_excel=ReadExcel()
if __name__ == '__main__':
    # print(do_excel.get_title())
    # print(do_excel.get_data())
    a=do_excel.get_data_obj()
    # do_excel.write_color()

    for i in a:
        print(i.title)




