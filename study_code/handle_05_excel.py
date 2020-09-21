# encoding:utf-8
# author:wangzhicheng
# time:2019/12/18 11:34
# file:handle_05_excel.py


from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors, PatternFill
from TestScripts.handle_01_path import do_path
from TestScripts.handle_02_yaml import do_yaml


class DataObj:
    """测试数据类"""
    pass


class HandleExcel:
    """处理excel中数据"""

    def __init__(self, excel_path=do_path.excel_path,sheet_name=do_yaml.read_yaml("excel", "sheet_name_1")):

        self.excel_path = excel_path
        self.sheet_name=sheet_name

    # 打开excel文件，把sheetname放这，方便后面调用
    def open_excel(self):
        # 定位文件，创建文件对象
        self.wb = load_workbook(self.excel_path)
        # 定位文件名
        self.sheet = self.wb[self.sheet_name]

    # 获取excel中title数据
    def get_title(self):

        # 调用方法，打开文件
        self.open_excel()
        # 列表推导式，获取title数据
        title = [i.value for i in list(self.sheet.rows)[0]]
        # 关闭文件
        self.wb.close()
        # 返回数据
        return title

    def get_datas(self):
        title = self.get_title()

        self.open_excel()

        datas = []

        for i in list(self.sheet.rows)[1:]:
            data = [j.value for j in i]

            data_obj = DataObj()

            for k in zip(title, data):
                setattr(data_obj, k[0], k[1])
            datas.append(data_obj)

        return datas

    def get_datas_1(self):

        title = self.get_title()

        self.open_excel()

        datas = [dict(zip(title, [j.value for j in i])) for i in list(self.sheet.rows)[1:]]

        # datas=[]
        # for i in list(self.sheet.rows)[1:]:
        #     data=[]
        #     for j in i:
        #         data.append(j.value)
        #     datas.append(dict(zip(title,data)))

        self.wb.close()

        return datas

    def write_datas(self, row, column, value):

        self.open_excel()

        # 指定单元格row,column，写入value值
        self.sheet.cell(row, column, value)
        # 写完之后保存文件
        self.wb.save(self.excel_path)
        # 保存之后关闭文件，释放内存
        self.wb.close()

    def write_color(self, title_name=do_yaml.read_yaml("excel", "title_name")):

        title = self.get_title()

        self.open_excel()

        # 设置单元格背景颜色
        style_1 = PatternFill(patternType="solid", fgColor=colors.RED)
        style_2 = PatternFill(patternType="solid", fgColor=colors.GREEN)
        style_3 = PatternFill(patternType="solid", fgColor=colors.WHITE)

        # 遍历 result这一列数据，如果单元格的值 ==result，则使用指定的单元格背景颜色
        # title是一个列表，所有标题在其中，result标题在列表中的i索引就是，
        # 按列将每列存放进一个列表中，对应到result这列额索引就是，result在title中的索引
        for i in list(self.sheet.columns)[title.index(title_name)]:
            if i.value == do_yaml.read_yaml("excel", "pass_name"):
                i.fill = style_2
            elif i.value == do_yaml.read_yaml("excel", "fail_name"):
                i.fill = style_1
            else:
                i.fill = style_3

        # 设置完后，保存文件
        self.wb.save(self.excel_path)
        self.wb.close()


do_excel = HandleExcel()

if __name__ == '__main__':

    # title = do_excel.get_title()
    # print(title)
    #
    # datas = do_excel.get_datas()
    #
    # for i in datas:
    #     print(i.title)
    #
    # datas_1 = do_excel.get_datas_1()
    # print(datas_1)
    do_excel = HandleExcel(sheet_name="recharge")

    data=do_excel.get_datas()

    for i in data:
        print(eval(i.datas))

